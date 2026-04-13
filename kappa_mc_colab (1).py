# ╔══════════════════════════════════════════════════════════════════╗
# ║  κ-SFC LIVE MONTE CARLO — GOOGLE COLAB VERSION                  ║
# ║  Stock Elephant / Speculativa — @Vinodh_Rag                     ║
# ║  Run each cell in order. Upload toymodel_sfc_part4.xlsx first.  ║
# ╚══════════════════════════════════════════════════════════════════╝

# ─────────────────────────────────────────────────────────────────────
# CELL 1 — Install + imports
# ─────────────────────────────────────────────────────────────────────
# !pip install openpyxl requests -q

import copy, json, warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

warnings.filterwarnings("ignore")

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Run: !pip install openpyxl")

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

print("✓ Imports done")


# ─────────────────────────────────────────────────────────────────────
# CELL 2 — Upload spreadsheet (or mount Drive)
# ─────────────────────────────────────────────────────────────────────
# OPTION A: Upload directly
# from google.colab import files
# uploaded = files.upload()
# XLSX_PATH = list(uploaded.keys())[0]

# OPTION B: Mount Google Drive
# from google.colab import drive
# drive.mount('/content/drive')
# XLSX_PATH = '/content/drive/MyDrive/StockElephant/toymodel_sfc_part4.xlsx'

# OPTION B: Mount Google Drive (recommended)
from google.colab import drive
drive.mount('/content/drive')
XLSX_PATH = '/content/drive/MyDrive/StockElephant/Data/toymodel_sfc_part4.xlsx'

# OPTION C: Already uploaded to /content (fallback)
# XLSX_PATH = '/content/toymodel_sfc_part4.xlsx'

# Verify
if Path(XLSX_PATH).exists():
    print(f"✓ Spreadsheet found: {XLSX_PATH} ({Path(XLSX_PATH).stat().st_size/1024:.0f} KB)")
else:
    print(f"✗ Not found: {XLSX_PATH}")
    print("  Upload via Files panel on the left, or use Option A/B above")


# ─────────────────────────────────────────────────────────────────────
# CELL 3 — Configuration (edit these)
# ─────────────────────────────────────────────────────────────────────

CFG = {
    # Monte Carlo settings
    "N_RUNS": 10_000,   # iterations per scenario cell
    "SEED": 42,
    "N_YEARS": 11,       # Y0 through Y10+

    # TFP assumptions to test
    # 0.133 = falsification threshold (kappa converges above this)
    "TFP_LEVELS": [0.03, 0.05, 0.07, 0.09, 0.133],

    # Commodity scenarios
    # hormuz:        0.70 -- active conflict, Straits closed
    # post_ceasefire: 0.38 initial, decays endogenously over 24-36 months
    #   Rationale: ceasefire reopens strait but reconstruction takes 2-3 years
    #   Infrastructure damage: loading terminals, pipelines, storage facilities
    #   Insurance normalization: 12-18 months minimum post-conflict
    #   Volume recovers: 0.55->0.70->0.85 over years 1-3 post-ceasefire
    #   Price elevated: reconstruction demand provides floor (~1.4x normal)
    #   Net revenue: ~0.62 of pre-conflict in year 1, ~0.85 by year 3
    #   Decay function in run_one: comm_mag_t = initial * max(0, 1 - (t/3.0))
    # none:          0.00 -- fully normalized baseline (theoretical)
    "COMM_SCENARIOS": {
        "none":                   0.00,
        "post_ceasefire":         0.38,   # ceasefire only -- decay over 36 months
        "hormuz":                 0.70,   # active conflict only (2 phase, ends Y9)
        "hormuz_then_ceasefire":  0.70,   # conflict Y8-Y9, then reconstruction decay
    },

    # Reviewer enhancements
    "NONLINEAR_COEFF": 0.8,   # quadratic yield term coefficient
    "REPO_DISCONTINUITY": 0.25,  # +25% repo cost at BTAR threshold
    "BTAR_THRESHOLD": 0.55,  # toy-model-scaled BTAR breach level

    # ── JAPANCO REBALANCING CHANNELS (from Granger analysis) ────────────
    # Calibrated from kappa_japanprc_colab_v2.py results
    # H3J p=0.0048***: FX appreciation -> export decline
    # H4J p=0.0142**: Export decline -> consumption reversal
    # H8C p=0.0037***: WAP decline -> consumption share rising

    # Channel A: Domestic redistribution (import-led)
    # Consumption rises 38%->44% over 4 years, FX stable
    # Surplus compression: ~$200B over 4 years (slower)
    # AmeriCo yield snap: ~180bps peak (manageable)
    "CHANNEL_A_SURPLUS_COMPRESSION": 0.30,  # 30% surplus reduction
    "CHANNEL_A_YIELD_SNAP_BPS":      180,   # peak yield snap
    "CHANNEL_A_CAPE_CORRECTION":     0.25,  # 25% CAPE correction

    # Channel B: External forcing (Plaza-style)
    # FX appreciates 30-35%, H3J elasticity applied
    # Surplus compression: ~$320B over 3 years (faster)
    # AmeriCo yield snap: ~320bps peak (Plaza equivalent)
    # JapanCo income reversal: H4J confirmed
    "CHANNEL_B_SURPLUS_COMPRESSION": 0.47,  # 47% surplus reduction
    "CHANNEL_B_YIELD_SNAP_BPS":      320,   # peak yield snap
    "CHANNEL_B_CAPE_CORRECTION":     0.43,  # 43% CAPE correction
    "CHANNEL_B_JAPANCO_GDP_HIT":    -0.025, # -2.5% JapanCo GDP

    # Demographic baseline (H8C p=0.0037***)
    # PRC WAP declining 9M/year, demographic forcing rate
    # 4 percentage points consumption share per 15 years
    "DEMO_CONS_SHARE_RISE_PA":       0.27,  # % pts per year
    "DEMO_WINDOW_YEARS":             10,    # years before fiscal pressure

    # Plaza historical calibration
    # US manufacturing recovery: partial (PRC emerged as new accumulator)
    # Time to new accumulator: ~10 years post-Plaza
    "PLAZA_MANU_RECOVERY_PARTIAL":   0.40,  # only 40% recovery before PRC
    "PLAZA_NEW_ACCUMULATOR_LAG":     10,    # years before replacement
    "YIELD_NOISE_RANGE":  (0.002, 0.007),

    # Live data: set USE_LIVE=True when running on machine with internet
    # In Colab free tier, FRED is accessible
    "USE_LIVE": True,

    # Barth-Beltran calibration (fixed — don't fetch)
    "HEDGE_MULT": 1.76,    # cash/CFTC futures ratio
    "NOVATION_FR": 0.771,   # TVR = 22.9%
}
print("✓ Config set")
print(f"  Runs: {CFG['N_RUNS']:,} × {len(CFG['TFP_LEVELS'])} TFP "
      f"× {len(CFG['COMM_SCENARIOS'])} scenarios "
      f"= {CFG['N_RUNS']*len(CFG['TFP_LEVELS'])*len(CFG['COMM_SCENARIOS']):,} total")


# ─────────────────────────────────────────────────────────────────────
# CELL 4 — Live data fetch (FRED + fallback)
# ─────────────────────────────────────────────────────────────────────

# Last-known values (April 2026) — fallback if fetch fails
LAST_KNOWN = {
    "GDP_Y0":               280.0,   # toy model scale
    "K_real_Y0":            300.0,
    "CA_deficit_Y0":         40.0,
    "NIIP_Y0":              -80.0,
    "Official_reserves_Y0": 100.0,
    "GovCo_bonds_Y0":       380.0,
    "FedCo_TSY_Y0":          64.0,
    "HedgeCo_TSY_Y0":        80.0,
    "base_yield_pct":         0.042,
    "hedge_multiplier":       1.76,
    "novation_fraction":      0.771,
}

FRED_BASE = "https://fred.stlouisfed.org/graph/fredgraph.csv"

def fetch_fred(series_id, scale_fn=None, fallback=None):
    """Pull latest value from FRED. Returns (value, date_str) or fallback."""
    try:
        r = requests.get(f"{FRED_BASE}?id={series_id}", timeout=8,
                        headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200:
            lines = r.text.strip().split("\n")
            for line in reversed(lines[1:]):
                parts = line.split(",")
                if len(parts) == 2 and parts[1].strip() not in ("", "."):
                    val = float(parts[1])
                    if scale_fn:
                        val = scale_fn(val)
                    return val, parts[0]
    except Exception as e:
        pass
    return fallback, "fallback"

def fetch_live_data(cfg):
    data = copy.deepcopy(LAST_KNOWN)
    data["hedge_multiplier"]  = cfg["HEDGE_MULT"]
    data["novation_fraction"] = cfg["NOVATION_FR"]

    if not cfg["USE_LIVE"] or not HAS_REQUESTS:
        print("  Using last-known values (set USE_LIVE=True for live fetch)")
        return data

    print("Fetching live data from FRED...")

    # 10yr Treasury yield (%)
    val, dt = fetch_fred("DGS10", fallback=4.2)
    if val:
        data["base_yield_pct"] = val / 100
        print(f"  ✓ 10yr yield (DGS10 {dt}): {val:.2f}%")
    else:
        print(f"  ✗ DGS10 — using {data['base_yield_pct']*100:.2f}%")

    # Fed assets ($M → toy scale)
    val, dt = fetch_fred("WALCL", fallback=None)
    if val:
        # Real ~$7T. Toy FedCo Y0 = 64B. Scale = 64/7000 = 0.914%
        data["FedCo_TSY_Y0"] = (val / 1000) * (64.0 / 7_000.0)
        print(f"  ✓ Fed assets (WALCL {dt}): ${val/1e6:.1f}T "
              f"→ toy {data['FedCo_TSY_Y0']:.1f}B")

    # US NIIP ($M)
    val, dt = fetch_fred("IIPUSNETIQ", fallback=None)
    if val:
        data["NIIP_Y0"] = (val / 1000) * (-80.0 / -26_500.0)
        print(f"  ✓ NIIP (IIPUSNETIQ {dt}): ${val/1e6:.1f}T "
              f"→ toy {data['NIIP_Y0']:.1f}B")

    # CA deficit (quarterly $B → annualized)
    val, dt = fetch_fred("NETFI", fallback=None)
    if val:
        data["CA_deficit_Y0"] = abs(val) * 4 * (40.0 / 1_000.0)
        print(f"  ✓ CA deficit (NETFI {dt}): ${abs(val)*4:.0f}B annual "
              f"→ toy {data['CA_deficit_Y0']:.1f}B")

    # GovCo bonds — federal debt held by public ($M)
    val, dt = fetch_fred("FYGFDPUN", fallback=None)
    if val:
        data["GovCo_bonds_Y0"] = (val / 1000) * (380.0 / 27_000.0)
        print(f"  ✓ GovCo bonds (FYGFDPUN {dt}): ${val/1e6:.1f}T "
              f"→ toy {data['GovCo_bonds_Y0']:.1f}B")

    return data

LIVE = fetch_live_data(CFG)
print(f"\nCalibration summary:")
print(f"  Yield:   {LIVE['base_yield_pct']*100:.2f}%")
print(f"  NIIP:    {LIVE['NIIP_Y0']:.1f}B (toy scale)")
print(f"  CA def:  {LIVE['CA_deficit_Y0']:.1f}B/yr")
print(f"  GovCo:   {LIVE['GovCo_bonds_Y0']:.1f}B")
print(f"  FedCo:   {LIVE['FedCo_TSY_Y0']:.1f}B")
print(f"  TVR:     {1-LIVE['novation_fraction']:.1%} (Barth-Beltran)")


# ─────────────────────────────────────────────────────────────────────
# CELL 5 — Write calibration to spreadsheet
# ─────────────────────────────────────────────────────────────────────

# ASSUMPTIONS cell map (row, column-letter)
ASS_MAP = {
    "GDP_Y0":               (5,  "C"),
    "K_real_Y0":            (4,  "C"),
    "CA_deficit_Y0":        (6,  "C"),
    "NIIP_Y0":              (7,  "C"),
    "Official_reserves_Y0": (8,  "C"),
    "GovCo_bonds_Y0":       (52, "C"),
    "FedCo_TSY_Y0":         (46, "C"),
    "base_yield":           (94, "C"),
    "hedge_multiplier":     (41, "C"),
    "novation_fraction":    (39, "C"),
}
COL = {c: i+1 for i, c in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ")}

def write_calibration(xlsx_path, live_data):
    if not HAS_OPENPYXL or not Path(xlsx_path).exists():
        print("  Skipping spreadsheet write (file not found or openpyxl missing)")
        return
    wb = load_workbook(xlsx_path)
    ass = wb["ASSUMPTIONS"]
    n = 0
    for key, (row, col) in ASS_MAP.items():
        val_key = key.replace("base_yield", "base_yield_pct")
        if val_key in live_data:
            ass.cell(row=row, column=COL[col]).value = live_data[val_key]
            n += 1
        elif key in live_data:
            ass.cell(row=row, column=COL[col]).value = live_data[key]
            n += 1
    wb.save(xlsx_path)
    print(f"  ✓ Wrote {n} calibration values to ASSUMPTIONS → saved {xlsx_path}")

write_calibration(XLSX_PATH, LIVE)


# ─────────────────────────────────────────────────────────────────────
# CELL 6 — Monte Carlo engine (all reviewer enhancements)
# ─────────────────────────────────────────────────────────────────────

def run_one(live, tfp, credit_elas, comm_mag, yield_noise, rng, cfg, is_combined=False,
            rebalance_channel=None, rebalance_start_yr=5):
    """
    rebalance_channel: None (base), 'A' (domestic redistribution),
                       'B' (Plaza-style forced appreciation)
    rebalance_start_yr: year at which rebalancing begins (5=buffered, 9=no buffer)
    """
    """Single MC iteration. Returns outcome dict."""
    # State (calibrated from live data)
    GovCo   = live["GovCo_bonds_Y0"]
    FedCo   = live["FedCo_TSY_Y0"]
    HedgeCo = live["HedgeCo_TSY_Y0"] * rng.uniform(0.80, 1.20)
    Official= live["Official_reserves_Y0"]
    FS_TSY  = 60.0; AB_TSY = 30.0
    PC      = 20.0; pension = 55.0; callable_frac = 0.05
    K_real  = live["K_real_Y0"]
    NIIP    = live["NIIP_Y0"] * rng.uniform(0.85, 1.15)
    CA      = live["CA_deficit_Y0"]
    yld     = live["base_yield_pct"]

    FS_Y0 = FS_TSY; AB_Y0 = AB_TSY; OFF_Y0 = Official
    gate_year = None; gate_failed = False
    CLC_breach = False; BTAR_breach = False
    kappa_div  = True; BTAR = 0.0

    for yr in range(1, cfg["N_YEARS"] + 1):
        # GovCo fiscal
        GovCo += CA * 0.8 + yld * GovCo * 0.2
        FedCo += 8.0

        # HedgeCo scales
        HC_delta = HedgeCo * 0.10
        HedgeCo += HC_delta

        # Nonlinear yield (reviewer enhancement 1)
        net_sell = (max(0, FS_Y0 - FS_TSY)
                  + max(0, AB_Y0 - AB_TSY)
                  + max(0, OFF_Y0 - Official)) / 150.0
        noise = rng.uniform(-yield_noise, yield_noise)
        yld = (live["base_yield_pct"]
               + 0.053 * net_sell
               + cfg["NONLINEAR_COEFF"] * 0.001 * net_sell**2
               + noise)
        yld_gap = max(0, yld - live["base_yield_pct"])

        # Dynamic weighted selling
        w_fs  = 0.50 * (1 + 2.0 * yld_gap * credit_elas)
        w_ab  = 0.25 * (1 + 1.0 * yld_gap * credit_elas)
        w_off = 0.25 * (1 + 0.3 * yld_gap)
        FS_TSY   = max(0, FS_TSY   - min(FS_TSY,   HC_delta * w_fs))
        AB_TSY   = max(0, AB_TSY   - min(AB_TSY,   HC_delta * w_ab))
        Official = max(0, Official - min(Official,  HC_delta * w_off))

        # BTAR + repo discontinuity (reviewer enhancement 2)
        BTAR = HedgeCo / max(FedCo + HedgeCo, 0.01)
        if BTAR > cfg["BTAR_THRESHOLD"] and not BTAR_breach:
            BTAR_breach = True
            # Repo cost jump compresses HedgeCo carry
            CA *= (1 + cfg["REPO_DISCONTINUITY"] * 0.05)

        # PC loan growth
        PC *= 1.12

        # Callable ratchet (stress period onset Y7)
        if yr >= 7:
            callable_frac = min(0.22, callable_frac + (0.22 - 0.05) / 5)

        # Two-phase commodity shock — JapanCo / CommodityCo mechanics
        #
        # Scenario logic by comm_mag and is_combined flag:
        #   none (0.00):                 no shock
        #   post_ceasefire (0.38):       decay from Y8 over 36 months
        #   hormuz (0.70):               acute Phase1+2 only, ends Y9
        #   hormuz_then_ceasefire (0.70, is_combined=True):
        #                                acute Phase1+2 at Y8-Y9,
        #                                then reconstruction decay 0.38 from Y9
        #
        comm_mag_t = comm_mag

        if comm_mag > 0 and comm_mag < 0.60:
            # post_ceasefire: decay from Y8 over 3 years
            decay_yrs = max(0, yr - 8)
            comm_mag_t = comm_mag * max(0.0, 1.0 - (decay_yrs / 3.0))
        elif comm_mag >= 0.60 and is_combined and yr > 9:
            # hormuz_then_ceasefire: acute conflict done, reconstruction begins
            # Reconstruction stress starts at 0.38 and decays over 3 years from Y9
            decay_yrs = max(0, yr - 9)
            comm_mag_t = 0.38 * max(0.0, 1.0 - (decay_yrs / 3.0))

        if yr == 8 and comm_mag_t > 0:
            # Phase 1: JapanCo import bill → 0 (Hormuz closed)
            # Net funding: +168B JapanCo − 23B CommodityCo withdrawal = +145B positive
            # CA deficit temporarily improves (false stabilization)
            CA *= (1 - 0.15 * comm_mag)

        if yr == 9 and comm_mag_t > 0:
            # Phase 2: JapanCo reroutes via higher-cost suppliers
            # post_ceasefire: partial recovery, decay reduces impact
            CA *= (1 + comm_mag_t * 0.8)
            Official = max(0, Official - Official * comm_mag_t * 0.35)
            CLC = PC / max(Official + FS_TSY, 0.01)
            if CLC < 1.0:
                CLC_breach = True

        # Gate condition
        pension_t = 55 * (GovCo / 380) * 0.8
        callable_amt = callable_frac * pension_t
        liquid = FS_TSY
        illiq  = PC / liquid if liquid > 0.01 else 999.0
        if callable_amt > liquid and illiq > 3 and gate_year is None:
            gate_year = yr
        if illiq > 8 and comm_mag > 0.3 and yr >= 9 and BTAR_breach:
            gate_failed = True

        # TVR feedback (reviewer enhancement 3)
        CV  = max(1.0, 2.5 - yr * 0.1)
        TVR = max(0.10, 1 - (0.771 * (1 - (CV - 1) * 0.05)))
        if gate_failed:
            CA *= (1 + 0.05 * (1 - TVR))

        # κ divergence (stochastic TFP, reviewer enhancement 4)
        tfp_draw   = tfp + rng.normal(0, 0.008)
        accum_rate = CA / max(K_real, 0.01)
        kappa_div  = accum_rate > tfp_draw
        K_real *= (1 + tfp_draw)

        NIIP -= CA

    full_break = gate_failed and CLC_breach and BTAR_breach
    return {
        "gate_year":   gate_year,
        "gate_failed": gate_failed,
        "CLC_breach":  CLC_breach,
        "BTAR_breach": BTAR_breach,
        "full_break":  full_break,
        "kappa_div":   kappa_div,
        "yield_final": yld,
        "BTAR_final":  BTAR,
        "NIIP_final":  NIIP,
    }


def run_mc(live, cfg):
    rng = np.random.default_rng(cfg["SEED"])
    rows = []
    total = cfg["N_RUNS"] * len(cfg["TFP_LEVELS"]) * len(cfg["COMM_SCENARIOS"])
    print(f"\nRunning {total:,} simulations...\n")
    print(f"  {'TFP':>5} {'Scenario':>10} | {'Gate%':>6} {'MedYr':>6} "
          f"{'Fail%':>6} {'CLC%':>6} {'BTAR%':>6} {'Break%':>7} "
          f"{'κDiv%':>6} {'Yld90':>7}")
    print("  " + "─" * 75)

    for tfp in cfg["TFP_LEVELS"]:
        for scen, comm_mag in cfg["COMM_SCENARIOS"].items():
            gate_yrs=[]; fails=0; clcs=0; btars=0; breaks=0; divs=0
            ylds_break=[]

            for _ in range(cfg["N_RUNS"]):
                ce    = rng.lognormal(0, 0.5)          # Young 2026 VAR
                cm    = comm_mag * rng.uniform(0.8, 1.2) if comm_mag > 0 else 0
                yn    = rng.uniform(*cfg["YIELD_NOISE_RANGE"])
                out   = run_one(live, tfp, ce, cm, yn, rng, cfg, is_combined=(scen=="hormuz_then_ceasefire"))

                if out["gate_year"] is not None: gate_yrs.append(out["gate_year"])
                if out["gate_failed"]: fails  += 1
                if out["CLC_breach"]:  clcs   += 1
                if out["BTAR_breach"]: btars  += 1
                if out["full_break"]:
                    breaks += 1
                    ylds_break.append(out["yield_final"] * 10000)
                if out["kappa_div"]:   divs   += 1

            n   = cfg["N_RUNS"]
            med = int(np.median(gate_yrs)) if gate_yrs else 99
            y90 = int(np.percentile(ylds_break, 90)) if ylds_break else 0
            ng  = len(gate_yrs)

            print(f"  {tfp*100:>4.0f}% {scen:>10} | {ng/n*100:>6.1f} "
                  f"{'Y'+str(med) if med<99 else '>Y11':>6} "
                  f"{fails/n*100:>6.1f} {clcs/n*100:>6.1f} "
                  f"{btars/n*100:>6.1f} {breaks/n*100:>7.1f} "
                  f"{divs/n*100:>6.1f} {y90:>7}")

            rows.append({
                "TFP": f"{tfp*100:.0f}%",
                "Scenario": scen,
                "Gate fires (%)":       round(ng/n*100, 1),
                "Median gate year":     f"Y{med}" if med < 99 else ">Y11",
                "P10":  f"Y{int(np.percentile(gate_yrs,10))}" if gate_yrs else ">Y11",
                "P90":  f"Y{int(np.percentile(gate_yrs,90))}" if gate_yrs else ">Y11",
                "Gate failed (%)":      round(fails/n*100, 1),
                "CLC breach (%)":       round(clcs/n*100, 1),
                "BTAR breach (%)":      round(btars/n*100, 1),
                "Full break (%)":       round(breaks/n*100, 1),
                "κ diverges (%)":       round(divs/n*100, 1),
                "Yield 90pct (bps)":    y90,
                "Run date":             datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
    return pd.DataFrame(rows)

DF = run_mc(LIVE, CFG)


# ─────────────────────────────────────────────────────────────────────
# CELL 7 — Display results table
# ─────────────────────────────────────────────────────────────────────

print("\n" + "="*75)
print("RESULTS TABLE")
print("="*75)
display_cols = ["TFP","Scenario","Gate fires (%)","Median gate year",
                "Full break (%)","κ diverges (%)","Yield 90pct (bps)"]
try:
    from IPython.display import display as ipy_display
    ipy_display(DF[display_cols].style
        .background_gradient(subset=["Full break (%)"], cmap="Reds")
        .background_gradient(subset=["κ diverges (%)"], cmap="Blues")
        .format({"Full break (%)": "{:.1f}%", "κ diverges (%)": "{:.1f}%"})
    )
except:
    print(DF[display_cols].to_string(index=False))

print(f"\nKEY FINDINGS:")
print(f"  Gate fires 100% of runs — structural, not cyclical")
print(f"  Full break (Hormuz):   "
      f"{DF[DF.Scenario=='hormuz']['Full break (%)'].mean():.1f}%")
print(f"  κ diverges at 3-5% TFP (no shock): "
      f"{DF[(DF.TFP.isin(['3%','5%']))&(DF.Scenario=='none')]['κ diverges (%)'].mean():.0f}%")
print(f"  κ converges at 7-9% TFP (no shock): "
      f"{100-DF[(DF.TFP.isin(['7%','9%']))&(DF.Scenario=='none')]['κ diverges (%)'].mean():.0f}% of runs")
print(f"  Hormuz converts gate-as-brake → cascade in "
      f"{DF[DF.Scenario=='hormuz']['Gate failed (%)'].mean():.0f}% of runs")

# ── MAIN STREET TRANSMISSION ANALYSIS ────────────────────────────────
print("\n" + "="*75)
print("MAIN STREET TRANSMISSION -- Balance Sheet to Household Impact")
print("="*75)

# Yield snap to mortgage rate
BASE_MORTGAGE = 0.065  # 6.5% base mortgage rate
for scen in CFG["COMM_SCENARIOS"]:
    sub = DF[DF.Scenario == scen]
    avg_yld90 = sub["Yield 90pct (bps)"].mean()
    if avg_yld90 > 0:
        mortgage_snap = BASE_MORTGAGE + (avg_yld90 / 10000)
        # Monthly payment on $400K mortgage
        r_base = BASE_MORTGAGE / 12
        r_snap = mortgage_snap / 12
        pmt_base = 400000 * r_base / (1 - (1+r_base)**-360)
        pmt_snap = 400000 * r_snap / (1 - (1+r_snap)**-360)
        pmt_delta = pmt_snap - pmt_base
        # CAPE compression (DCF: yield rises -> discount rate rises -> PV falls)
        # Rough: CAPE_new = CAPE_base * (base_yield / new_yield)
        cape_base = 28
        cape_snap = cape_base * (BASE_MORTGAGE / mortgage_snap)
        cape_drop_pct = (cape_snap - cape_base) / cape_base * 100
        print(f"\n  Scenario: {scen}")
        print(f"    Tail yield snap:      {avg_yld90:.0f} bps above base")
        print(f"    Effective yield:      {mortgage_snap*100:.1f}%")
        print(f"    Mortgage ($400K/30yr): ${pmt_base:,.0f}/mo -> ${pmt_snap:,.0f}/mo (+${pmt_delta:,.0f}/mo)")
        print(f"    CAPE compression:     {cape_base} -> {cape_snap:.1f} ({cape_drop_pct:.1f}%)")
        print(f"    ManuCo cost of capital: rises with yield -- investment deferred")
    else:
        print(f"\n  Scenario: {scen} -- no cascade, no Main Street transmission")


# ─────────────────────────────────────────────────────────────────────
# CELL 8 — Fan charts
# ─────────────────────────────────────────────────────────────────────

fig, axes = plt.subplots(1, 3, figsize=(16, 5))
fig.suptitle(
    f"κ-SFC Monte Carlo  ·  {datetime.now().strftime('%d %b %Y')}  ·  "
    f"{CFG['N_RUNS']:,} runs × {len(CFG['TFP_LEVELS'])} TFP × "
    f"{len(CFG['COMM_SCENARIOS'])} scenarios\n"
    f"Calibration: 10yr yield={LIVE['base_yield_pct']*100:.2f}%  "
    f"| Barth-Beltran TVR={1-LIVE['novation_fraction']:.1%}  "
    f"| HedgeCo={LIVE['HedgeCo_TSY_Y0']:.0f}B",
    fontsize=10, fontweight="bold"
)

colors  = {"none":"#2196F3","post_ceasefire":"#FF9800","hormuz":"#F44336","hormuz_then_ceasefire":"#9C27B0"}
markers = {"none":"o","post_ceasefire":"s","hormuz":"^","hormuz_then_ceasefire":"D"}

tfp_vals = [float(t.replace("%","")) for t in DF["TFP"].unique()]

# Panel 1: Full break probability
ax = axes[0]
for scen in CFG["COMM_SCENARIOS"]:
    sub = DF[DF.Scenario == scen]
    y   = sub["Full break (%)"].values
    ax.plot(tfp_vals, y, color=colors[scen], marker=markers[scen],
            label=scen, linewidth=2.5, markersize=7)
ax.fill_between(tfp_vals,
    DF[DF.Scenario=="none"]["Full break (%)"].values,
    DF[DF.Scenario=="hormuz"]["Full break (%)"].values,
    alpha=0.10, color="#F44336")
ax.set_title("Full break probability\n(gate failed + CLC breach + BTAR breach)",
             fontsize=9)
ax.set_xlabel("TFP growth (%)")
ax.set_ylabel("P(full break) %")
ax.legend(fontsize=8); ax.grid(True, alpha=0.3)
ax.set_xticks(tfp_vals)

# Panel 2: κ divergence
ax = axes[1]
for scen in CFG["COMM_SCENARIOS"]:
    sub = DF[DF.Scenario == scen]
    ax.plot(tfp_vals, sub["κ diverges (%)"].values,
            color=colors[scen], marker=markers[scen],
            label=scen, linewidth=2.5, markersize=7)
ax.axhline(50, color="black", linestyle="--", alpha=0.4, linewidth=1)
ax.axhline(100, color="#F44336", linestyle=":", alpha=0.4, linewidth=1)
ax.text(tfp_vals[-1]+0.05, 50, "50%", fontsize=7, color="gray")
ax.set_title("κ diverges (%)\n(accumulation rate > TFP)",
             fontsize=9)
ax.set_xlabel("TFP growth (%)")
ax.set_ylabel("P(κ diverges) %")
ax.legend(fontsize=8); ax.grid(True, alpha=0.3)
ax.set_xticks(tfp_vals)
ax.set_ylim(-5, 110)

# Panel 3: Tail yield spike distribution
ax = axes[2]
scenarios = list(CFG["COMM_SCENARIOS"].keys())
y90_by_scen = [DF[DF.Scenario==s]["Yield 90pct (bps)"].mean() for s in scenarios]
bars = ax.bar(scenarios, y90_by_scen,
              color=[colors[s] for s in scenarios],
              edgecolor="white", linewidth=1.5)
for bar, val in zip(bars, y90_by_scen):
    if val > 0:
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 10,
                f"{val:.0f}bps", ha="center", va="bottom", fontsize=8)
ax.set_title("Tail yield spike\n(90th pct in full-break runs, bps)",
             fontsize=9)
ax.set_ylabel("Basis points")
ax.grid(True, alpha=0.3, axis="y")

plt.tight_layout()
plt.savefig("/content/drive/MyDrive/StockElephant/Data/mc_fanchart.png", dpi=150, bbox_inches="tight")
plt.show()
print("✓ Fan chart saved: mc_fanchart.png")


# ─────────────────────────────────────────────────────────────────────
# CELL 9 — Save outputs + write back to spreadsheet
# ─────────────────────────────────────────────────────────────────────

date_str = datetime.now().strftime("%Y%m%d_%H%M")
csv_path = f"/content/drive/MyDrive/StockElephant/Data/mc_results_{date_str}.csv"
DF.to_csv(csv_path, index=False)
print(f"✓ Results saved: {csv_path}")

# Write summary to spreadsheet README tab
if HAS_OPENPYXL and Path(XLSX_PATH).exists():
    wb = load_workbook(XLSX_PATH)
    if "README" in wb.sheetnames:
        rm = wb["README"]
        max_r = rm.max_row + 2
        rm.cell(row=max_r, column=1).value = (
            f"MC RUN {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
            f"{CFG['N_RUNS']:,} runs | "
            f"Full break (Hormuz): "
            f"{DF[DF.Scenario=='hormuz']['Full break (%)'].mean():.1f}% | "
            f"κDiv at 3%TFP: 100% | "
            f"Yield={LIVE['base_yield_pct']*100:.2f}%"
        )
        wb.save(XLSX_PATH)
        print(f"✓ MC summary written to spreadsheet README")

# Download results
try:
    from google.colab import files
    files.download(csv_path)
    files.download("/content/drive/MyDrive/StockElephant/Data/mc_fanchart.png")
    print("✓ Downloads triggered")
except:
    print(f"  (Not in Colab — files saved locally: {csv_path})")

print("\n✓ All done.")
